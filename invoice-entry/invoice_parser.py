"""
Invoice Parser
Extracts key fields from PDF invoices and outputs a spreadsheet for fast manual entry.

Modes:
    Outlook mode (Windows):  python invoice_parser.py --outlook
    Folder mode:             python invoice_parser.py [folder_path]

Outlook mode pulls unread emails with PDF attachments from Inbox > CBS Invoices,
parses each invoice, marks the email complete (green checkmark), and marks it read.
"""

import anthropic
import base64
import json
import os
import re
import sys
import tempfile
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL = True
except ImportError:
    EXCEL = False

OUTLOOK_FOLDER = "CBS Invoices"


# ── PDF extraction via Claude ─────────────────────────────────────────────────

EXTRACT_PROMPT = """\
Look at this document and return a JSON object with the following fields.

First, determine what type of document this is:
- "invoice" = a bill requesting payment for specific goods/services
- "statement" = a summary of account activity or balance over a period
- "other" = anything else

If the document type is NOT "invoice", return only:
{"document_type": "statement"}

If it IS an invoice, extract these fields:
- document_type: "invoice"
- vendor_name: Company name of who is billing (not the recipient)
- po_number: The PO number, job number, or order reference field, exactly as printed
- invoice_number: The invoice number or invoice ID
- invoice_date: Invoice date in MM/DD/YY format
- invoice_total: The final total amount owed as a plain number (no $ or commas)

Return ONLY valid JSON, no explanation. Example:
{"document_type":"invoice","vendor_name":"ABC Supply","po_number":"CP26006-17","invoice_number":"18366029-00","invoice_date":"4/13/26","invoice_total":"83.48"}
"""

def extract_from_pdf(pdf_path, client):
    with open(pdf_path, "rb") as f:
        pdf_b64 = base64.standard_b64encode(f.read()).decode("utf-8")

    response = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=300,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "document",
                    "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": pdf_b64,
                    },
                    "cache_control": {"type": "ephemeral"},
                },
                {
                    "type": "text",
                    "text": EXTRACT_PROMPT,
                },
            ],
        }],
    )

    raw = response.content[0].text.strip()
    raw = re.sub(r"^```[a-z]*\n?", "", raw)
    raw = re.sub(r"\n?```$", "", raw)
    return json.loads(raw)


# ── Outlook integration (Windows only) ───────────────────────────────────────

def get_pdfs_from_outlook():
    """
    Connects to Outlook, finds unread emails with PDF attachments in
    Inbox > CBS Invoices, saves PDFs to a temp folder, marks each email
    complete (green checkmark) and read.

    Returns: (temp_dir, list of (pdf_path, fname)) tuples
    """
    try:
        import win32com.client
    except ImportError:
        print("ERROR: pywin32 not installed. Run: pip install pywin32")
        sys.exit(1)

    print("Connecting to Outlook...")
    outlook = win32com.client.Dispatch("Outlook.Application")
    namespace = outlook.GetNamespace("MAPI")

    inbox = namespace.GetDefaultFolder(6)  # 6 = olFolderInbox
    try:
        folder = inbox.Folders[OUTLOOK_FOLDER]
    except Exception:
        print(f"ERROR: Could not find folder 'Inbox > {OUTLOOK_FOLDER}'")
        print("Check that the folder name matches exactly.")
        sys.exit(1)

    temp_dir = Path(tempfile.mkdtemp(prefix="invoices_"))
    results = []

    items = folder.Items
    total = items.Count
    unread_count = 0

    print(f"Scanning '{OUTLOOK_FOLDER}' ({total} emails)...")

    for i in range(1, total + 1):
        try:
            item = items[i]
        except Exception:
            continue

        if item.Class != 43:  # 43 = olMail
            continue
        if not item.UnRead:
            continue

        pdf_found = False
        for attachment in item.Attachments:
            fname = attachment.FileName
            if fname.lower().endswith(".pdf"):
                save_path = temp_dir / fname
                if save_path.exists():
                    save_path = temp_dir / f"{save_path.stem}_{i}.pdf"
                attachment.SaveAsFile(str(save_path))
                results.append((save_path, fname))
                pdf_found = True
                unread_count += 1

        if pdf_found:
            item.FlagStatus = 1
            item.TaskCompletedDate = datetime.now()
            item.UnRead = False
            item.Save()

    if not results:
        print(f"No unread emails with PDF attachments found in '{OUTLOOK_FOLDER}'.")
        sys.exit(0)

    print(f"Found {unread_count} unread invoice(s).\n")
    return temp_dir, results


# ── Excel output ──────────────────────────────────────────────────────────────

HEADERS = ["Vendor", "PO #", "Invoice #", "Date", "Total", "File", "Date Parsed"]

def update_excel(new_rows, output_dir):
    timestamp = datetime.now().strftime("%Y-%m-%d %I-%M %p")
    output_path = output_dir / f"Invoice List - {timestamp}.xlsx"

    existing_rows = []
    old_files = sorted(output_dir.glob("Invoice List - *.xlsx"))

    if old_files:
        old_file = old_files[-1]
        wb_existing = openpyxl.load_workbook(old_file)
        ws_existing = wb_existing.active
        for row in ws_existing.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                row_list = list(row)
                while len(row_list) < len(HEADERS):
                    row_list.append("")
                existing_rows.append(row_list[:len(HEADERS)])
        old_file.unlink()

    all_rows = existing_rows + new_rows
    all_rows.sort(key=lambda r: (str(r[0] or "").lower(), str(r[1] or "").lower()))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(all_rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 5:  # Total
                cell.number_format = '"$"#,##0.00'
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="D6E4F0")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(output_path)
    return output_path


# ── Core processing ───────────────────────────────────────────────────────────

def process_pdfs(pdf_entries, client):
    rows = []
    errors = []

    for pdf_path, display_name in pdf_entries:
        print(f"  Processing: {display_name} ...", end=" ", flush=True)
        try:
            data = extract_from_pdf(pdf_path, client)

            if data.get("document_type") != "invoice":
                print("SKIPPED (statement)")
                continue

            po_raw = data.get("po_number", "")

            total = data.get("invoice_total", "")
            try:
                total = float(str(total).replace(",", ""))
            except (ValueError, AttributeError):
                pass

            rows.append([
                data.get("vendor_name", ""),
                po_raw,
                data.get("invoice_number", ""),
                data.get("invoice_date", ""),
                total,
                display_name,
                datetime.now().strftime("%m/%d/%y %I:%M %p"),
            ])
            print("OK")
        except Exception as e:
            print(f"FAILED ({e})")
            errors.append((display_name, str(e)))

    return rows, errors


def print_table(rows):
    col_widths = [max(len(str(r[i])) for r in [HEADERS] + rows) for i in range(len(HEADERS))]
    divider = "+-" + "-+-".join("-" * w for w in col_widths) + "-+"

    def fmt(r):
        return "| " + " | ".join(str(r[i]).ljust(col_widths[i]) for i in range(len(HEADERS))) + " |"

    print(f"\n{divider}")
    print(fmt(HEADERS))
    print(divider)
    for row in rows:
        print(fmt(row))
    print(divider)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("ERROR: ANTHROPIC_API_KEY environment variable not set.")
        print("Set it in Windows: System Properties > Environment Variables")
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)
    outlook_mode = "--outlook" in sys.argv

    if outlook_mode:
        temp_dir, email_results = get_pdfs_from_outlook()
        pdf_entries = [(path, fname) for path, fname in email_results]
        output_dir = Path(r"C:\Users\NathanPayne.DESKTOP-UAVOJRP\OneDrive - MIRC Construction Services\Documents\3. CBS Legal Docs\AP Parser")
        output_dir.mkdir(parents=True, exist_ok=True)
    else:
        folder = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(".")
        pdfs = sorted(folder.glob("*.pdf"))
        if not pdfs:
            print(f"No PDF files found in: {folder.resolve()}")
            sys.exit(1)
        print(f"Found {len(pdfs)} PDF(s) in {folder.resolve()}\n")
        pdf_entries = [(p, p.name) for p in pdfs]
        output_dir = folder
        temp_dir = None

    rows, errors = process_pdfs(pdf_entries, client)

    if not rows:
        print("\nNo invoices were successfully parsed.")
        sys.exit(1)

    print_table(rows)

    if errors:
        print(f"\nFailed to parse {len(errors)} file(s):")
        for name, err in errors:
            print(f"  {name}: {err}")

    if EXCEL:
        output_path = update_excel(rows, output_dir)
        print(f"\nSaved to: {output_path}")

    if temp_dir and temp_dir.exists():
        import shutil
        shutil.rmtree(temp_dir)


if __name__ == "__main__":
    main()
